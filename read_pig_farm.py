#!/usr/bin/env python3

import sys
from pathlib import Path

try:
	import pandas as pd  # type: ignore
except Exception:  # pylint: disable=broad-except
	pd = None  # type: ignore


def _read_with_pandas(excel_path: Path) -> None:
	excel_file = pd.ExcelFile(excel_path, engine="openpyxl")
	sheet_names = excel_file.sheet_names
	print(f"Found sheets: {sheet_names}")

	first_sheet_name = sheet_names[0]
	dataframe = pd.read_excel(excel_file, sheet_name=first_sheet_name)
	print(f"Loaded sheet: {first_sheet_name} with shape {tuple(dataframe.shape)}")

	print("Columns and dtypes:")
	for column_name, dtype in zip(dataframe.columns, dataframe.dtypes):
		print(f"  - {column_name}: {dtype}")

	print("\nPreview (first 10 rows):")
	print(dataframe.head(10).to_string(index=False))


def _read_with_openpyxl(excel_path: Path) -> None:
	from datetime import timedelta
	from openpyxl import load_workbook  # type: ignore
	from openpyxl.utils.datetime import from_excel as excel_from_serial  # type: ignore

	workbook = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
	sheet_names = workbook.sheetnames
	print(f"Found sheets: {sheet_names}")

	first_sheet_name = sheet_names[0]
	sheet = workbook[first_sheet_name]
	print(f"Loaded sheet: {first_sheet_name}")

	rows_iter = sheet.iter_rows(values_only=True)
	try:
		header = next(rows_iter)
	except StopIteration:
		print("Sheet is empty")
		return

	print("Columns:")
	for column_index, column_name in enumerate(header or [], start=1):
		print(f"  - c{column_index}: {column_name}")

	# Try to locate Date and Time columns
	date_idx = None
	time_idx = None
	if header is not None:
		for idx, name in enumerate(header):
			name_str = ("" if name is None else str(name)).strip().lower()
			if name_str == "date":
				date_idx = idx
			elif name_str == "time":
				time_idx = idx

	print("\nPreview (first 10 data rows):")
	row_count = 0
	for row in rows_iter:
		cells = list(row)

		# Build a human-readable timestamp if possible
		timestamp_display = None
		try:
			if date_idx is not None and time_idx is not None:
				date_val = cells[date_idx]
				time_val = cells[time_idx]
				if isinstance(date_val, (int, float)) and isinstance(time_val, (int, float)):
					date_dt = excel_from_serial(date_val, epoch=workbook.epoch)
					time_td = excel_from_serial(time_val, epoch=workbook.epoch, timedelta=True)
					timestamp_display = (date_dt + time_td).strftime("%Y-%m-%d %H:%M")
		except Exception:
			pass

		# Print the row with optional timestamp prefix
		row_text = "\t".join("" if cell is None else str(cell) for cell in cells)
		if timestamp_display is not None:
			print(f"{timestamp_display}\t{row_text}")
		else:
			print(row_text)

		row_count += 1
		if row_count >= 10:
			break


def main() -> None:
	"""Load the Excel file, display sheet names, schema, and a small preview."""
	excel_path = Path("/workspace/pig_farm_7_2024.xlsx")
	if not excel_path.exists():
		print(f"Error: file not found: {excel_path}", file=sys.stderr)
		sys.exit(1)

	try:
		if pd is not None:
			_read_with_pandas(excel_path)
		else:
			_read_with_openpyxl(excel_path)
	except Exception as exc:  # pylint: disable=broad-except
		print("Failed to read Excel:", file=sys.stderr)
		print(str(exc), file=sys.stderr)
		sys.exit(2)


if __name__ == "__main__":
	main()