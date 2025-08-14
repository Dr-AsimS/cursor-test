#!/usr/bin/env python3

from __future__ import annotations

import math
import sys
from typing import Dict, List, Optional, Tuple
from pathlib import Path

from openpyxl import load_workbook  # type: ignore


def is_number(value: object) -> bool:
	return isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value))


def pearson_correlation(x_values: List[float], y_values: List[float]) -> Optional[float]:
	if len(x_values) != len(y_values) or len(x_values) < 2:
		return None
	mean_x = sum(x_values) / len(x_values)
	mean_y = sum(y_values) / len(y_values)
	numerator = 0.0
	sum_sq_x = 0.0
	sum_sq_y = 0.0
	for x_val, y_val in zip(x_values, y_values):
		dx = x_val - mean_x
		dy = y_val - mean_y
		numerator += dx * dy
		sum_sq_x += dx * dx
		sum_sq_y += dy * dy
	denominator = math.sqrt(sum_sq_x * sum_sq_y)
	if denominator == 0:
		return None
	return numerator / denominator


def main() -> None:
	excel_path = Path("/workspace/pig_farm_7_2024.xlsx")
	if not excel_path.exists():
		print(f"Error: file not found: {excel_path}", file=sys.stderr)
		sys.exit(1)

	workbook = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
	sheet = workbook[workbook.sheetnames[0]]

	rows_iter = sheet.iter_rows(values_only=True)
	try:
		header = list(next(rows_iter))
	except StopIteration:
		print("Sheet is empty", file=sys.stderr)
		sys.exit(2)

	# Build column index map
	name_to_index: Dict[str, int] = {}
	for idx, name in enumerate(header):
		if name is None:
			continue
		name_to_index[str(name).strip()] = idx

	# Expected column names
	indoor_col = None
	for candidate in ("Indoor-Temp(°C)", "Indoor-Temp", "Indoor Temp", "Indoor Temperature"):
		if candidate in name_to_index:
			indoor_col = candidate
			break
	if indoor_col is None:
		print("Could not find 'Indoor-Temp(°C)' column", file=sys.stderr)
		sys.exit(3)

	candidate_numeric_columns = []
	for col_name in ("Date", "Time", "Humidity(%)", "CO2(ppm)"):
		if col_name in name_to_index:
			candidate_numeric_columns.append(col_name)

	# Collect data pairs for correlation
	correlations: List[Tuple[str, Optional[float], int]] = []
	for other_col in candidate_numeric_columns:
		x_vals: List[float] = []  # Indoor temp
		y_vals: List[float] = []  # Other variable
		for row in rows_iter:
			# rows_iter is consumed; so we must re-create it per column
			pass
		# Re-create the iterator for each variable
		rows_iter_inner = sheet.iter_rows(values_only=True)
		try:
			next(rows_iter_inner)  # skip header
		except StopIteration:
			break
		for row in rows_iter_inner:
			indoor_val = row[name_to_index[indoor_col]] if name_to_index[indoor_col] < len(row) else None
			other_val = row[name_to_index[other_col]] if name_to_index[other_col] < len(row) else None
			if is_number(indoor_val) and is_number(other_val):
				x_vals.append(float(indoor_val))
				y_vals.append(float(other_val))
		corr = pearson_correlation(x_vals, y_vals)
		correlations.append((other_col, corr, len(x_vals)))

	print(f"Sheet: {sheet.title}")
	print(f"Target: {indoor_col}")
	print("Pearson correlation with numeric variables:")
	for var_name, corr, n in correlations:
		if corr is None:
			print(f"  - {var_name}: insufficient or constant data (n={n})")
		else:
			print(f"  - {var_name}: r = {corr:.4f} (n={n})")

	print("\nNote: Categorical variables like 'Weather' and 'Control' are excluded. They require encoding (e.g., one-hot) or a different association measure (e.g., correlation ratio).")


if __name__ == "__main__":
	main()